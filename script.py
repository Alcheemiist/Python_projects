import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl import Workbook
from sqlalchemy import create_engine
from openpyxl import load_workbook

# Connection to your database
engine = create_engine('postgresql://postgres:mysecretpassword@localhost:5432/postgres')

# # Connection to your database
# conn = psycopg2.connect(
#     dbname="postgres",  # replace with your database name
#     user="postgres",  # replace with your username
#     password="mysecretpassword",  # replace with your password
#     host="localhost",  # replace with your host
#     port="5432"  # replace with your host
# )

print('Connected to database')

# Check if the output file exists, and create it if not
try:
    book = load_workbook('./output.xlsx')
except FileNotFoundError:
    book = Workbook()
    book.save('./output.xlsx')
    book = load_workbook('./output.xlsx')

with pd.ExcelWriter('./output.xlsx', engine='openpyxl') as writer:
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    for script, sheet in zip(sql_scripts, excel_sheets):
        with open(script, 'r') as file:
            sql = file.read()  # read the SQL query from file
            data = pd.read_sql(sql, engine)  # pass the input_id as a parameter
            data.to_excel(writer, sheet_name=sheet, index=False)

    writer.save()
print('Disconnected from database')